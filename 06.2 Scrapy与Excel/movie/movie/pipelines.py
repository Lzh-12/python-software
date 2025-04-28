import logging
# from movie.items import MovieItem
from items import MovieItem
import openpyxl  # pip install openpyxl
import openpyxl.drawing.image  # pip install pillow


class MoviePipeline:
    def __init__(self):
        logging.info('将爬取到的电影数据保存在Excel文件之中。')
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = '豆瓣电影排名'
        self.worksheet.append(('电影排名', '电影名称', '发行区域', '发行时间',
                               '电影类型', '演员数量', '演员列表', '评分', '投票数量', '电影图片'))
        self.row = 2  # 记录行数

    def process_item(self, item: MovieItem, spider):
        self.worksheet.cell(row=self.row, column=1, value=item['rank'])
        self.worksheet.cell(row=self.row, column=2, value=item['title'])
        self.worksheet.cell(row=self.row, column=3, value=','.join(item['regions']))
        self.worksheet.cell(row=self.row, column=4, value=item['release_date'])
        self.worksheet.cell(row=self.row, column=5, value=','.join(item['types']))
        self.worksheet.cell(row=self.row, column=6, value=item['actor_count'])
        self.worksheet.cell(row=self.row, column=7, value=','.join(item['actors']))
        self.worksheet.cell(row=self.row, column=8, value=item['score'])
        self.worksheet.cell(row=self.row, column=9, value=item['vote_count'])
        image = openpyxl.drawing.image.Image(item['cover'])
        image.width = 170
        image.height = 251
        image.anchor = ('J' + str(self.row))  # 将图片锚定在单元格
        image.horizontal = 'center'  # 设置图片水平居中
        image.vertical = 'center'  # 设置图片垂直居中
        self.worksheet.add_image(image, ('J' + str(self.row)))
        self.worksheet.column_dimensions['J'].width = 30  # 设置A列的宽度
        self.worksheet.row_dimensions[self.row].height = 200  # 设置第1行的高度
        self.row += 1
        return item

    def close_spider(self, spider):
        logging.info('爬虫操作完成，数据保存到Excel文件之中。')
        self.workbook.save('movie.xlsx')
